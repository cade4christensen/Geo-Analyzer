#!/usr/bin/env python3
"""
Merge successful rows from a retry-analyzed file back into the original
analyzed file at their original row positions.

Usage:
    python3 merge_retry.py batch_input_analyzed.xlsx batch_retry_analyzed.xlsx

After this runs, batch_input_analyzed.xlsx will have the retry successes
overwritten in place; FAILED rows that are still FAILED after retry stay
as FAILED in the original (with the latest error message).
"""
import sys
from pathlib import Path

import openpyxl


def main():
    if len(sys.argv) != 3:
        print("Usage: python3 merge_retry.py <original_analyzed.xlsx> <retry_analyzed.xlsx>",
              file=sys.stderr)
        sys.exit(1)

    orig_path  = Path(sys.argv[1])
    retry_path = Path(sys.argv[2])
    if not orig_path.exists():
        print(f"Original analyzed file not found: {orig_path}", file=sys.stderr)
        sys.exit(1)
    if not retry_path.exists():
        print(f"Retry analyzed file not found: {retry_path}", file=sys.stderr)
        sys.exit(1)

    # Open the original analyzed workbook (writable).
    wb_orig = openpyxl.load_workbook(orig_path)
    ws_orig = wb_orig.active
    orig_headers = [c.value for c in next(ws_orig.iter_rows(min_row=1, max_row=1))]
    orig_summary_col = orig_headers.index("Nearest Feature (m)") + 1

    # Open the retry analyzed workbook (read-only).
    wb_retry = openpyxl.load_workbook(retry_path, data_only=True)
    ws_retry = wb_retry.active
    retry_headers = [c.value for c in next(ws_retry.iter_rows(min_row=1, max_row=1))]

    # Find "Source Row" column in the retry file.
    if "Source Row" not in retry_headers:
        print("Retry file is missing the 'Source Row' column — can't merge.",
              file=sys.stderr)
        sys.exit(1)
    src_row_col = retry_headers.index("Source Row") + 1
    retry_summary_col = retry_headers.index("Nearest Feature (m)") + 1

    # The retry file's analyzed columns start where the original 4 input cols
    # + "Source Row" end. Map: retry column N (for N >= retry_summary_col) →
    # original column N - 1 (the "Source Row" column shifts everything left
    # by 1 in the destination).
    # Actually: both files have analyzed columns starting at "Nearest Feature (m)"
    # so we map by header name.

    # Build a mapping from analyzed header → (orig_col, retry_col)
    analyzed_headers = [h for h in orig_headers
                        if isinstance(h, str) and (
                            h.startswith("Nearest ") or h in ("Surface", "Population"))]
    mapping = []
    for h in analyzed_headers:
        if h in retry_headers:
            mapping.append((h, orig_headers.index(h) + 1, retry_headers.index(h) + 1))

    print(f"Mapping {len(mapping)} analyzed columns from retry → original")
    print(f"Retry rows to process: {ws_retry.max_row - 1}")

    promoted = 0
    still_failed = 0
    skipped = 0

    for r in range(2, ws_retry.max_row + 1):
        src_row = ws_retry.cell(row=r, column=src_row_col).value
        if src_row is None:
            skipped += 1
            continue
        try:
            src_row = int(src_row)
        except (TypeError, ValueError):
            skipped += 1
            continue

        new_summary = ws_retry.cell(row=r, column=retry_summary_col).value
        if isinstance(new_summary, str) and new_summary.startswith(("FAILED:", "OUTSIDE_US", "BAD_COORDS")):
            still_failed += 1
            # Update the original's tag to the latest one so the user sees
            # the freshest error message.
            ws_orig.cell(row=src_row, column=orig_summary_col, value=new_summary)
            continue

        # Successful retry — copy every analyzed column back.
        for _, orig_c, retry_c in mapping:
            ws_orig.cell(row=src_row, column=orig_c,
                         value=ws_retry.cell(row=r, column=retry_c).value)
        promoted += 1
        print(f"  row {src_row}: retry succeeded → merged into original")

    wb_orig.save(orig_path)
    print(f"\nDone. Merged into {orig_path.name}:")
    print(f"  promoted to success: {promoted}")
    print(f"  still failed:        {still_failed}")
    print(f"  skipped (bad rows):  {skipped}")


if __name__ == "__main__":
    main()
