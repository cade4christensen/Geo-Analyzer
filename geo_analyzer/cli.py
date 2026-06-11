#!/usr/bin/env python3
"""
Command-line interface for the coordinate analyzer.

Usage examples:
  python cli.py 40.7128 -74.0060
  python cli.py --lat 37.7749 --lon -122.4194
  python cli.py --batch coords.txt
  python cli.py --json 40.7128 -74.0060
  python cli.py --excel input.xlsx
"""

import argparse
import json
import os
import sys
import time
from typing import Optional
from analyzer import analyze


# ---------------------------------------------------------------------------
# Rounding helper
# ---------------------------------------------------------------------------

def _round_dist(value, round_to: str):
    """Round a distance value according to round_to ('whole' or 'hundredth')."""
    if not isinstance(value, (int, float)):
        return value
    if round_to == "whole":
        return int(round(value, 0))
    return round(value, 2)


# ---------------------------------------------------------------------------
# Text formatting
# ---------------------------------------------------------------------------

def format_results(result: dict, round_to: str = "hundredth") -> str:
    lines = []
    c = result["coordinate"]
    lines.append(f"\n=== Coordinate Analysis: {c['lat']}, {c['lon']} ===\n")

    lines.append("LINEAR FEATURES (distance to nearest):")
    lf = result["linear_features"]
    for label in ("Road", "Trail", "Railroad", "Utility Line", "Water Feature"):
        info = lf.get(label, {})
        dist = info.get("distance_m")
        name = info.get("name")
        if dist is not None:
            dist = _round_dist(dist, round_to)
            fmt = ".0f" if round_to == "whole" else ".2f"
            name_str = f'  ({name})' if name else ""
            lines.append(f"  {label:<15} {dist:>10{fmt}} m{name_str}")
        else:
            lines.append(f"  {label:<15}     none found within 2 km")

    lines.append(f"\nSURFACE AT POINT:  {result['surface']}")
    lines.append(f"POPULATION:        {result['population']}")
    lines.append("")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Excel processing
# ---------------------------------------------------------------------------

# Always-visible summary columns (shown when detail group is collapsed)
SUMMARY_COLUMNS = [
    "Nearest Feature (m)",
    "Nearest Feature Type",
]

# Detail columns grouped and collapsible (one pair per feature type)
FEATURE_KEYS = ("Road", "Trail", "Railroad", "Utility Line", "Water Feature")

DETAIL_COLUMNS = []
for _fk in FEATURE_KEYS:
    DETAIL_COLUMNS.append(f"Nearest {_fk} (m)")
    DETAIL_COLUMNS.append(f"Nearest {_fk} Name")

# Final always-visible columns after the detail group
TAIL_COLUMNS = ["Surface", "Population"]


def _find_column(headers: list, *candidates: str) -> Optional[int]:
    """Return 0-based column index for the first matching header (case-insensitive)."""
    lower = [str(h).strip().lower() if h is not None else "" for h in headers]
    for c in candidates:
        if c.lower() in lower:
            return lower.index(c.lower())
    return None


def process_excel(input_path: str, round_to: str = "hundredth") -> str:
    """
    Read input_path (.xlsx), analyze every coordinate row, write results
    to a new file with '_analyzed' suffix. Returns the output path.

    Column layout:
      [input cols] | Nearest Feature (m) | Nearest Feature Type
                   | <-- grouped/collapsible detail columns -->
                   | Nearest Road (m) | Road Name | Trail (m) | ...
                   | Surface | Population
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("Error: openpyxl not installed. Run: pip3 install openpyxl", file=sys.stderr)
        sys.exit(1)

    # data_only=True reads cached formula results (e.g. =LEFT(B,FIND(",",B)-1)*1)
    # instead of the formula text itself.
    wb = openpyxl.load_workbook(input_path, data_only=True)
    ws = wb.active

    # --- Detect lat/lon columns ---
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    lat_col = _find_column(headers, "lat", "latitude")
    lon_col = _find_column(headers, "lon", "long", "longitude")
    # Optional fallback column with raw "lat,lon" text — used when lat/lon
    # cells are uncached formulas.
    coord_col = _find_column(headers, "coordinates", "coord", "coords",
                             "latlon", "lat,lon", "find.coord", "find coord")

    if lat_col is None or lon_col is None:
        print(
            "Error: could not find lat/lon columns.\n"
            "Expected headers named 'lat'/'latitude' and 'lon'/'longitude' (case-insensitive).",
            file=sys.stderr,
        )
        sys.exit(1)

    # --- Column positions ---
    # summary_start: Nearest Feature (m), Nearest Feature Type
    # detail_start:  per-feature distance + name pairs  (grouped)
    # tail_start:    Surface, Population
    summary_start = ws.max_column + 1
    detail_start  = summary_start + len(SUMMARY_COLUMNS)
    tail_start    = detail_start + len(DETAIL_COLUMNS)

    summary_fill = PatternFill("solid", fgColor="1F4E79")   # dark blue — always visible
    detail_fill  = PatternFill("solid", fgColor="2E75B6")   # mid blue  — detail group
    tail_fill    = PatternFill("solid", fgColor="1F4E79")   # dark blue — always visible
    hdr_font     = Font(bold=True, color="FFFFFF")
    center       = Alignment(horizontal="center")

    def write_header(col, name, fill):
        c = ws.cell(row=1, column=col, value=name)
        c.fill = fill
        c.font = hdr_font
        c.alignment = center

    # Summary headers
    for i, name in enumerate(SUMMARY_COLUMNS):
        write_header(summary_start + i, name, summary_fill)

    # Detail headers
    for i, name in enumerate(DETAIL_COLUMNS):
        write_header(detail_start + i, name, detail_fill)

    # Tail headers
    for i, name in enumerate(TAIL_COLUMNS):
        write_header(tail_start + i, name, tail_fill)

    # --- First pass: parse coords, analyze each row, mark failures ERROR ---
    total_rows = ws.max_row - 1
    # Track rows that need analysis (parseable coords) for use by retry loop.
    # row_jobs maps ws row number → (lat, lon).
    row_jobs: dict = {}
    # row_attempts counts how many times we've tried this row in retries.
    row_attempts: dict = {}
    # row_last_err remembers the last exception message per row (for the summary).
    row_last_err: dict = {}

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=1):
        ws_row = row_idx + 1
        lat_val = row[lat_col].value
        lon_val = row[lon_col].value

        # Try the lat/lon columns first.
        lat_f = lon_f = None
        try:
            if lat_val is not None and lon_val is not None:
                lat_f, lon_f = float(lat_val), float(lon_val)
        except (TypeError, ValueError):
            lat_f = lon_f = None

        # Fallback: parse "lat,lon" from the Coordinates column if the lat/lon
        # cells were uncached formulas or otherwise unreadable.
        if (lat_f is None or lon_f is None) and coord_col is not None:
            raw = row[coord_col].value
            if raw is not None:
                text = str(raw).replace("−", "-").strip()
                for sep in (",", " ", "\t"):
                    parts = text.split(sep, 1)
                    if len(parts) == 2:
                        try:
                            a, b = float(parts[0].strip()), float(parts[1].strip())
                            if -90 <= a <= 90 and -180 <= b <= 180:
                                lat_f, lon_f = a, b
                                break
                        except ValueError:
                            continue

        if lat_f is None or lon_f is None:
            # Bad coords — permanent failure, mark and skip retries.
            print(f"  Row {ws_row}: skipped (invalid lat/lon: {lat_val}, {lon_val})")
            _write_review_row(ws, ws_row, summary_start, detail_start, tail_start,
                              reason="BAD_COORDS")
            continue

        # Detect coordinates outside Lower-48 US — NLCD won't have data there,
        # so retries are pointless. L48 box: lat 24.4–49.4, lon −125 to −66.9.
        if not (24.4 <= lat_f <= 49.4 and -125.0 <= lon_f <= -66.9):
            print(f"  Row {ws_row}: outside US Lower 48 — skipping ({lat_f}, {lon_f})")
            _write_review_row(ws, ws_row, summary_start, detail_start, tail_start,
                              reason="OUTSIDE_US")
            continue

        # Eligible for analysis — record for the pass loop below.
        row_jobs[ws_row] = (lat_f, lon_f)
        row_attempts[ws_row] = 0

    # --- Analysis with auto-retry: up to MAX_PASSES, MAX_ATTEMPTS_PER_ROW ---
    MAX_PASSES = 5
    MAX_ATTEMPTS_PER_ROW = 3
    INTER_PASS_PAUSE_S = 8

    pending = set(row_jobs.keys())
    for pass_num in range(1, MAX_PASSES + 1):
        if not pending:
            break

        if pass_num == 1:
            print(f"\n=== Pass 1: analyzing {len(pending)} rows ===")
        else:
            print(f"\n=== Pass {pass_num}: retrying {len(pending)} ERROR rows "
                  f"(waiting {INTER_PASS_PAUSE_S}s for rate limits to cool) ===")
            time.sleep(INTER_PASS_PAUSE_S)

        completed = set()
        # Iterate pending in row order so progress reads top→bottom.
        for ws_row in sorted(pending):
            lat_f, lon_f = row_jobs[ws_row]
            row_attempts[ws_row] += 1
            attempt = row_attempts[ws_row]

            print(f"  [pass {pass_num}, row {ws_row}, attempt {attempt}] "
                  f"({lat_f}, {lon_f})...")
            try:
                result = analyze(lat_f, lon_f)
                _write_result_row(ws, ws_row, summary_start, detail_start, tail_start,
                                  result, round_to)
                completed.add(ws_row)
            except Exception as e:
                msg = str(e)[:200]
                row_last_err[ws_row] = msg
                print(f"    Error: {msg}")
                # Keep existing ERROR cells in place; will retry next pass
                _write_result_row(ws, ws_row, summary_start, detail_start, tail_start,
                                  None, round_to)
                # If this row has hit its attempt cap, give up and mark for review.
                if attempt >= MAX_ATTEMPTS_PER_ROW:
                    _write_review_row(ws, ws_row, summary_start, detail_start,
                                      tail_start, reason=f"FAILED:{msg[:60]}")
                    completed.add(ws_row)  # remove from pending; retries exhausted

        pending -= completed

    # --- Summary ---
    total_input = total_rows
    success_count = total_input - len(pending) - sum(
        1 for r in range(2, ws.max_row + 1)
        if ws.cell(row=r, column=summary_start).value in (
            "BAD_COORDS", "OUTSIDE_US")
        or (isinstance(ws.cell(row=r, column=summary_start).value, str)
            and ws.cell(row=r, column=summary_start).value.startswith("FAILED:"))
    )
    print(f"\n=== Batch complete ===")
    print(f"  Successful: {success_count} / {total_input}")
    if pending:
        print(f"  Permanently failed (still ERROR after {MAX_PASSES} passes): {len(pending)}")
        for ws_row in sorted(pending)[:10]:
            lat_f, lon_f = row_jobs[ws_row]
            err = row_last_err.get(ws_row, "no error captured")
            print(f"    row {ws_row} ({lat_f}, {lon_f}): {err[:100]}")
        if len(pending) > 10:
            print(f"    ...and {len(pending) - 10} more")

    # --- Auto-size columns ---
    all_cols = SUMMARY_COLUMNS + DETAIL_COLUMNS + TAIL_COLUMNS
    for i, name in enumerate(all_cols):
        letter = openpyxl.utils.get_column_letter(summary_start + i)
        ws.column_dimensions[letter].width = max(len(name) + 2, 16)

    # --- Group detail columns (collapsible, collapsed by default) ---
    # openpyxl's .group() only sets the first column, so set each one manually.
    for c in range(detail_start, tail_start):
        letter = openpyxl.utils.get_column_letter(c)
        ws.column_dimensions[letter].outlineLevel = 1
        ws.column_dimensions[letter].hidden = True

    ws.sheet_view.showOutlineSymbols = True

    # --- Save ---
    base, ext = os.path.splitext(input_path)
    output_path = base + "_analyzed" + ext
    wb.save(output_path)
    return output_path


def _write_review_row(ws, row: int, summary_start: int, detail_start: int,
                      tail_start: int, reason: str):
    """
    Mark a row with a permanent failure tag (instead of "ERROR") so the user
    can tell structural failures from transient ones. Used for: BAD_COORDS,
    OUTSIDE_US, FAILED:<exception>.
    """
    total_cols = len(SUMMARY_COLUMNS) + len(DETAIL_COLUMNS) + len(TAIL_COLUMNS)
    # First cell carries the tag, the rest stay blank so the spreadsheet stays
    # readable on review.
    ws.cell(row=row, column=summary_start, value=reason)
    for i in range(1, total_cols):
        ws.cell(row=row, column=summary_start + i, value="")


def _write_result_row(ws, row: int, summary_start: int, detail_start: int,
                      tail_start: int, result, round_to: str = "hundredth"):
    """Write one row of analysis results across all output column sections."""
    if result is None:
        for i in range(len(SUMMARY_COLUMNS) + len(DETAIL_COLUMNS) + len(TAIL_COLUMNS)):
            ws.cell(row=row, column=summary_start + i, value="ERROR")
        return

    lf = result["linear_features"]

    # Find the nearest feature overall for the summary columns
    nearest_dist = None
    nearest_type = None
    for key in FEATURE_KEYS:
        d = lf.get(key, {}).get("distance_m")
        if d is not None and (nearest_dist is None or d < nearest_dist):
            nearest_dist = d
            nearest_type = key

    ws.cell(row=row, column=summary_start,     value=_round_dist(nearest_dist, round_to) if nearest_dist is not None else "none within 2km")
    ws.cell(row=row, column=summary_start + 1, value=nearest_type or "none within 2km")

    # Detail columns
    col = detail_start
    for key in FEATURE_KEYS:
        info = lf.get(key, {})
        dist = info.get("distance_m")
        name = info.get("name")
        ws.cell(row=row, column=col,     value=_round_dist(dist, round_to) if dist is not None else "none within 2km")
        ws.cell(row=row, column=col + 1, value=name or "")
        col += 2

    # Tail columns
    ws.cell(row=row, column=tail_start,     value=result["surface"])
    ws.cell(row=row, column=tail_start + 1, value=result["population"])


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Analyze geographic coordinates using OpenStreetMap and NLCD data."
    )
    parser.add_argument("lat", type=float, nargs="?", help="Latitude (decimal degrees)")
    parser.add_argument("lon", type=float, nargs="?", help="Longitude (decimal degrees)")
    parser.add_argument("--lat", type=float, dest="lat_flag", metavar="LAT")
    parser.add_argument("--lon", type=float, dest="lon_flag", metavar="LON")
    parser.add_argument(
        "--batch", metavar="FILE",
        help="Text file with one 'lat,lon' or 'lat lon' per line"
    )
    parser.add_argument(
        "--excel", metavar="FILE",
        help="Excel (.xlsx) file with lat/lon columns; outputs FILE_analyzed.xlsx"
    )
    parser.add_argument(
        "--json", action="store_true",
        help="Output results as JSON instead of formatted text"
    )
    parser.add_argument(
        "--round", dest="round_to", choices=["whole", "hundredth"], default="hundredth",
        metavar="PRECISION",
        help="Distance rounding: 'whole' (nearest meter) or 'hundredth' (nearest 0.01 m). Default: hundredth"
    )

    args = parser.parse_args()
    round_to = args.round_to

    # --- Excel mode ---
    if args.excel:
        if not os.path.isfile(args.excel):
            print(f"Error: file '{args.excel}' not found.", file=sys.stderr)
            sys.exit(1)
        print(f"Processing {args.excel}...")
        output = process_excel(args.excel, round_to=round_to)
        print(f"\nDone. Results saved to: {output}")
        return

    # --- Single coordinate or batch text file ---
    lat = args.lat if args.lat is not None else args.lat_flag
    lon = args.lon if args.lon is not None else args.lon_flag
    coords = []

    if args.batch:
        try:
            with open(args.batch) as f:
                for line in f:
                    line = line.strip()
                    if not line or line.startswith("#"):
                        continue
                    parts = line.replace(",", " ").split()
                    if len(parts) >= 2:
                        coords.append((float(parts[0]), float(parts[1])))
        except FileNotFoundError:
            print(f"Error: file '{args.batch}' not found.", file=sys.stderr)
            sys.exit(1)
        except ValueError as e:
            print(f"Error parsing batch file: {e}", file=sys.stderr)
            sys.exit(1)
    elif lat is not None and lon is not None:
        coords.append((lat, lon))
    else:
        # No arguments — launch interactive mode
        _interactive_mode(round_to)
        return

    all_results = []
    for lat_i, lon_i in coords:
        result = analyze(lat_i, lon_i)
        all_results.append(result)
        if not args.json:
            print(format_results(result, round_to=round_to))

    if args.json:
        print(json.dumps(all_results if len(all_results) > 1 else all_results[0], indent=2))


def _interactive_mode(round_to: str = "hundredth"):
    print("=" * 50)
    print("  Coordinate Analyzer — Interactive Mode")
    print("  Type a coordinate to analyze, or 'q' to quit.")
    print("  Format: lat lon  or  lat,lon")
    print("  Example: 40.7128 -74.0060")
    print("=" * 50)

    while True:
        try:
            raw = input("\nEnter coordinate: ").strip()
        except (EOFError, KeyboardInterrupt):
            print("\nGoodbye.")
            break

        if raw.lower() in ("q", "quit", "exit"):
            print("Goodbye.")
            break

        parts = raw.replace(",", " ").split()
        if len(parts) < 2:
            print("  Please enter both latitude and longitude.")
            continue

        try:
            lat_i = float(parts[0])
            lon_i = float(parts[1])
        except ValueError:
            print("  Invalid numbers. Try again.")
            continue

        try:
            result = analyze(lat_i, lon_i)
            print(format_results(result, round_to=round_to))
        except Exception as e:
            print(f"  Error: {e}")


if __name__ == "__main__":
    main()
