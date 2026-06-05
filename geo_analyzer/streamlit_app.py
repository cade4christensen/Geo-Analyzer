import base64
import io
import time
from pathlib import Path
import streamlit as st
from analyzer import analyze

st.set_page_config(
    page_title="Coordinate Analyzer",
    page_icon="🗺️",
    layout="centered",
)

# ---------------------------------------------------------------------------
# Background — faint world map, slightly darker via pseudo-element opacity
# ---------------------------------------------------------------------------

_map_path = Path(__file__).parent / "static" / "world_map.png"
if _map_path.exists():
    _map_b64 = base64.b64encode(_map_path.read_bytes()).decode()
    st.markdown(f"""
    <style>
    /* Render the map in a pseudo-element so opacity only affects the image,
       not the page content. 0.18 = slightly darker than before (was ~0.12). */
    .stApp::before {{
        content: "";
        position: fixed;
        inset: 0;
        background-image: url("data:image/png;base64,{_map_b64}");
        background-size: cover;
        background-position: center;
        background-repeat: no-repeat;
        opacity: 0.12;
        z-index: 0;
        pointer-events: none;
    }}
    </style>
    """, unsafe_allow_html=True)

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _round_dist(value, decimals: int):
    """Round a distance value to `decimals` decimal places (0 = whole number)."""
    if not isinstance(value, (int, float)):
        return value
    rounded = round(value, decimals)
    return int(rounded) if decimals == 0 else rounded


def fmt_dist(value, decimals: int) -> str:
    if value is None:
        return "none within 2 km"
    v = _round_dist(value, decimals)
    if decimals == 0:
        return f"{v} m"
    return f"{v:.{decimals}f} m"


FEATURE_KEYS = ("Road", "Trail", "Railroad", "Utility Line", "Water Feature")

POPULATION_COLOR = {
    "Wilderness": "#4ade80",
    "Rural":      "#facc15",
    "Suburban":   "#fb923c",
    "Urban":      "#f87171",
}

SURFACE_COLOR = {
    "Structure": "#94a3b8",
    "Road":      "#64748b",
    "Linear":    "#a78bfa",
    "Drainage":  "#38bdf8",
    "Water":     "#0ea5e9",
    "Woods":     "#22c55e",
    "Scrub":     "#84cc16",
    "Field":     "#eab308",
    "Rock":      "#d1d5db",
    "Unknown":   "#6b7280",
}


def badge(label: str, color: str) -> str:
    return (
        f'<span style="background:{color};color:#000;padding:3px 10px;'
        f'border-radius:12px;font-weight:600;font-size:0.85rem;">{label}</span>'
    )


def display_results(result: dict, decimals: int):
    lf = result["linear_features"]
    surface = result["surface"]
    population = result["population"]

    sc = SURFACE_COLOR.get(surface, "#6b7280")
    pc = POPULATION_COLOR.get(population, "#6b7280")
    st.markdown(
        f"**Surface:** {badge(surface, sc)}&nbsp;&nbsp;"
        f"**Population:** {badge(population, pc)}",
        unsafe_allow_html=True,
    )
    st.markdown("")

    nearest_dist = None
    nearest_type = None
    for key in FEATURE_KEYS:
        d = lf.get(key, {}).get("distance_m")
        if d is not None and (nearest_dist is None or d < nearest_dist):
            nearest_dist = d
            nearest_type = key

    if nearest_type:
        st.info(f"**Nearest feature:** {nearest_type} — {fmt_dist(nearest_dist, decimals)}")

    st.markdown("#### Linear Features")
    rows = []
    for key in FEATURE_KEYS:
        info = lf.get(key, {})
        dist = info.get("distance_m")
        name = info.get("name") or ""
        source = info.get("source")
        dist_str = fmt_dist(dist, decimals)
        # Tag the cell when a non-default data source was used (e.g. TIGER fallback).
        if source and source != "OSM" and dist is not None:
            dist_str = f"{dist_str} ({source})"
        rows.append({
            "Feature":    key,
            "Distance":   dist_str,
            "Name / Ref": name,
        })
    st.table(rows)

    # Debug info: surfaces the signals that drove the population decision
    pop_debug = result.get("_pop_debug") or {}
    if pop_debug:
        with st.expander("Population signal debug"):
            st.json(pop_debug)

    # Per-source cascade results for the Road category (TIGER/USFS/USGS)
    road_cascade = result.get("_road_cascade") or []
    if road_cascade:
        with st.expander("Road cascade debug"):
            st.json(road_cascade)


# ---------------------------------------------------------------------------
# Excel processing
# ---------------------------------------------------------------------------

def process_excel_bytes(file_bytes: bytes, decimals: int) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    SUMMARY_COLUMNS = ["Nearest Feature (m)", "Nearest Feature Type"]
    DETAIL_COLUMNS = []
    for fk in FEATURE_KEYS:
        DETAIL_COLUMNS.append(f"Nearest {fk} (m)")
        DETAIL_COLUMNS.append(f"Nearest {fk} Name")
    TAIL_COLUMNS = ["Surface", "Population"]

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    lower = [str(h).strip().lower() if h else "" for h in headers]

    def find_col(*candidates):
        for c in candidates:
            if c.lower() in lower:
                return lower.index(c.lower())
        return None

    lat_col = find_col("lat", "latitude")
    lon_col = find_col("lon", "long", "longitude")

    if lat_col is None or lon_col is None:
        st.error("Could not find lat/lon columns.")
        return None

    summary_start = ws.max_column + 1
    detail_start  = summary_start + len(SUMMARY_COLUMNS)
    tail_start    = detail_start + len(DETAIL_COLUMNS)

    summary_fill = PatternFill("solid", fgColor="1F4E79")
    detail_fill  = PatternFill("solid", fgColor="2E75B6")
    hdr_font     = Font(bold=True, color="FFFFFF")
    center       = Alignment(horizontal="center")

    for i, name in enumerate(SUMMARY_COLUMNS):
        c = ws.cell(row=1, column=summary_start + i, value=name)
        c.fill = summary_fill; c.font = hdr_font; c.alignment = center
    for i, name in enumerate(DETAIL_COLUMNS):
        c = ws.cell(row=1, column=detail_start + i, value=name)
        c.fill = detail_fill; c.font = hdr_font; c.alignment = center
    for i, name in enumerate(TAIL_COLUMNS):
        c = ws.cell(row=1, column=tail_start + i, value=name)
        c.fill = summary_fill; c.font = hdr_font; c.alignment = center

    total = ws.max_row - 1
    progress = st.progress(0, text="Analyzing rows...")

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=1):
        progress.progress(row_idx / total, text=f"Analyzing row {row_idx} of {total}...")
        lat_val = row[lat_col].value
        lon_val = row[lon_col].value

        def write_error():
            for i in range(len(SUMMARY_COLUMNS) + len(DETAIL_COLUMNS) + len(TAIL_COLUMNS)):
                ws.cell(row=row_idx + 1, column=summary_start + i, value="ERROR")

        if lat_val is None or lon_val is None:
            write_error(); continue
        try:
            lat_f, lon_f = float(lat_val), float(lon_val)
        except (TypeError, ValueError):
            write_error(); continue

        try:
            if row_idx > 1:
                time.sleep(1)
            result = analyze(lat_f, lon_f)
        except Exception:
            write_error(); continue

        lf = result["linear_features"]
        nearest_dist = None
        nearest_type = None
        for key in FEATURE_KEYS:
            d = lf.get(key, {}).get("distance_m")
            if d is not None and (nearest_dist is None or d < nearest_dist):
                nearest_dist = d; nearest_type = key

        nd = _round_dist(nearest_dist, decimals) if nearest_dist is not None else "none within 2km"
        ws.cell(row=row_idx + 1, column=summary_start,     value=nd)
        ws.cell(row=row_idx + 1, column=summary_start + 1, value=nearest_type or "none within 2km")

        col = detail_start
        for key in FEATURE_KEYS:
            info = lf.get(key, {})
            dist = info.get("distance_m")
            name = info.get("name")
            ws.cell(row=row_idx + 1, column=col,     value=_round_dist(dist, decimals) if dist is not None else "none within 2km")
            ws.cell(row=row_idx + 1, column=col + 1, value=name or "")
            col += 2

        ws.cell(row=row_idx + 1, column=tail_start,     value=result["surface"])
        ws.cell(row=row_idx + 1, column=tail_start + 1, value=result["population"])

    all_cols = SUMMARY_COLUMNS + DETAIL_COLUMNS + TAIL_COLUMNS
    for i, name in enumerate(all_cols):
        letter = openpyxl.utils.get_column_letter(summary_start + i)
        ws.column_dimensions[letter].width = max(len(name) + 2, 16)

    for c in range(detail_start, tail_start):
        letter = openpyxl.utils.get_column_letter(c)
        ws.column_dimensions[letter].outlineLevel = 1
        ws.column_dimensions[letter].hidden = True
    ws.sheet_view.showOutlineSymbols = True

    progress.empty()
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ---------------------------------------------------------------------------
# Coordinate input helpers
# ---------------------------------------------------------------------------

def _try_parse_pair(text: str):
    """Return (lat, lon) floats if text looks like 'x,y' or 'x y', else None.

    Also normalises unicode minus signs and rejects out-of-range values so a
    typo can't produce a pyproj error downstream.
    """
    text = text.strip()
    # Normalise common unicode dashes that copy-paste can introduce
    text = text.replace("−", "-").replace("–", "-").replace("—", "-")
    for sep in (",", " ", "\t"):
        parts = text.split(sep, 1)
        if len(parts) == 2:
            try:
                a = float(parts[0].strip())
                b = float(parts[1].strip())
            except ValueError:
                continue
            if -90.0 <= a <= 90.0 and -180.0 <= b <= 180.0:
                return a, b
    return None


def coordinate_input():
    """
    Smart coordinate input.
    - Default: two side-by-side text inputs for lat and lon.
    - Typing a comma in either field auto-switches to combined mode.
    - The ⇌ button manually toggles between modes.
    Returns (lat, lon) floats or (None, None).
    """
    if "coord_split" not in st.session_state:
        st.session_state.coord_split = True
    if "combined_str" not in st.session_state:
        st.session_state.combined_str = ""

    input_col, btn_col = st.columns([11, 1])

    with btn_col:
        st.markdown("<div style='margin-top:28px'></div>", unsafe_allow_html=True)
        if st.button("⇌", help="Toggle between combined 'lat, lon' input and split inputs"):
            st.session_state.coord_split = not st.session_state.coord_split
            st.rerun()

    with input_col:
        if st.session_state.coord_split:
            c1, c2 = st.columns(2)
            with c1:
                lat_raw = st.text_input("Latitude", value="",
                                        placeholder="e.g. 40.7128", key="lat_txt")
            with c2:
                lon_raw = st.text_input("Longitude", value="",
                                        placeholder="e.g. -74.0060", key="lon_txt")

            # If the user pasted a full "lat, lon" pair into one field and left
            # the other empty, accept it in place without switching modes.
            if lat_raw.strip() and not lon_raw.strip():
                parsed = _try_parse_pair(lat_raw)
                if parsed:
                    return parsed
            if lon_raw.strip() and not lat_raw.strip():
                parsed = _try_parse_pair(lon_raw)
                if parsed:
                    return parsed

            # Auto-switch to combined mode only if BOTH fields have content and
            # one contains a comma — that's a clear sign the user is mid-paste
            # into the wrong layout.
            if lat_raw.strip() and lon_raw.strip():
                for raw in (lat_raw, lon_raw):
                    if "," in raw:
                        st.session_state.combined_str = raw
                        st.session_state.coord_split = False
                        st.rerun()

            # Parse the two separate fields (with bounds checking)
            lat, lon = None, None
            if lat_raw.strip():
                try:
                    v = float(lat_raw.strip().replace("−", "-"))
                    if -90.0 <= v <= 90.0:
                        lat = v
                    else:
                        st.caption("⚠️ Latitude out of range (-90 to 90)")
                except ValueError:
                    st.caption("⚠️ Invalid latitude")
            if lon_raw.strip():
                try:
                    v = float(lon_raw.strip().replace("−", "-"))
                    if -180.0 <= v <= 180.0:
                        lon = v
                    else:
                        st.caption("⚠️ Longitude out of range (-180 to 180)")
                except ValueError:
                    st.caption("⚠️ Invalid longitude")
            return lat, lon

        else:
            raw = st.text_input(
                "Coordinate  (lat, lon)",
                value=st.session_state.combined_str,
                placeholder="Paste e.g.  40.7128, -74.0060",
                key="combined_input",
            )
            st.session_state.combined_str = raw
            parsed = _try_parse_pair(raw) if raw else None
            if raw and parsed is None:
                st.caption("⚠️ Couldn't parse — use format: lat, lon")
            return (parsed[0], parsed[1]) if parsed else (None, None)


# ---------------------------------------------------------------------------
# Rounding control
# ---------------------------------------------------------------------------

def rounding_control() -> int:
    """
    Returns the number of decimal places to round to (0 = whole meter).
    UI: checkbox 'Round to whole meter'. If unchecked, a compact slider
    appears inline letting you pick 1–3 decimal places.
    """
    col_check, col_slider = st.columns([2, 3])
    with col_check:
        whole = st.checkbox("Round to nearest meter", value=False)
    if whole:
        return 0
    with col_slider:
        decimals = st.select_slider(
            "Decimal places",
            options=[1, 2, 3],
            value=2,
            format_func=lambda x: f"0.{'0'*(x-1)}1 m",
            label_visibility="collapsed",
        )
    return decimals


# ---------------------------------------------------------------------------
# UI
# ---------------------------------------------------------------------------

st.title("Coordinate Analyzer")
st.caption("Identifies linear features, surface type, and population density for any US coordinate.")

decimals = rounding_control()

st.markdown("---")
tab_single, tab_excel = st.tabs(["Single Coordinate", "Excel Batch"])

# ---- Single coordinate tab ----
with tab_single:
    lat, lon = coordinate_input()

    if st.button("Analyze", type="primary", use_container_width=True):
        if lat is None or lon is None:
            st.warning("Enter both a latitude and longitude.")
        else:
            with st.spinner("Querying OpenStreetMap and NLCD..."):
                try:
                    result = analyze(float(lat), float(lon))
                    st.success(f"Results for **{lat}, {lon}**")
                    display_results(result, decimals)
                except Exception as e:
                    st.error(f"Analysis failed: {e}")

# ---- Excel batch tab ----
with tab_excel:
    st.markdown(
        "Upload an `.xlsx` file with **`lat`** and **`lon`** columns. "
        "Results are appended as new columns and returned as a download."
    )
    uploaded = st.file_uploader("Choose an Excel file", type=["xlsx"])

    if uploaded is not None:
        if st.button("Run Analysis", type="primary", use_container_width=True):
            file_bytes = uploaded.read()
            result_bytes = process_excel_bytes(file_bytes, decimals)
            if result_bytes:
                st.success("Done!")
                st.download_button(
                    label="Download analyzed.xlsx",
                    data=result_bytes,
                    file_name=uploaded.name.replace(".xlsx", "_analyzed.xlsx"),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
